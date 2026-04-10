import re
import pdfplumber
from typing import Optional

# ── Patrones reales ───────────────────────────────────────────────────────────

# Número de factura — 6 variantes reales
PATRONES_FACTURA = [
    r"FACTURAS?\s+(?:DE\s+)?EXPORTACION\s*[.:/]\s*([\d,\s]+)",
    r"FACTURA\s+COMERCIAL\s*[.:/]\s*([\d,\s]+)",
    r"COMMERCIAL\s+INVOICE\s*[.:/]?\s*(?:N[°º])?\s*([\d,\s/]+)",
    r"INVOICE\s+N[°º]?\s*[.:/]\s*([\d,\s/]+)",
    r"N[°º]\s+FACTURA\s*[.:/]\s*([\d,\s]+)",
    r"FACT(?:URA)?\s*[.:/]\s*(\d+)",
]

# Blumar tiene formato especial: "139067/67254" (factura_chile/factura_cliente)
PATRON_FACTURA_BLUMAR = r"(\d{6,})\s*/\s*(\d{4,})"

# Incoterm — CPT es el más frecuente (260 casos), luego CFR (128), CIF (97)
PATRONES_INCOTERM = [
    r"CL[AÁ]USULA\s+DE\s+VENTA\s*[.:/]\s*([A-Z]{2,5})",
    r"INCOTERMS?\s*[.:/]\s*([A-Z]{2,5})",
    r"CONDICI[OÓ]N\s+DE\s+VENTA\s*[.:/]\s*([A-Z]{2,5})",
    r"\b(CPT|CFR|CIF|FOB|FCA|DAP|EXW|DDP)\b",
]

# Moneda
PATRONES_MONEDA = [
    r"TIPO\s+DE\s+MONEDA\s*[.:/]\s*([A-Z]{3})",
    r"CURRENCY\s*[.:/]\s*([A-Z]{3})",
    r"\b(USD|CNY|EUR|CLP)\b",
]

# Valor total
PATRONES_TOTAL = [
    r"TOTAL\s+[A-Z]{2,5}\s*:\s*([\d.,]+)",          # "TOTAL CPT : 109.648,80"
    r"TOTAL\s+(?:GENERAL\s+)?(?:USD|US\$|CLP|CNY|EUR)?\s*[.:/]?\s*([\d.,]+)",
    r"(?:USD|US\$)\s*([\d.,]+)\s*(?:TOTAL|$)",
    r"MONTO\s+TOTAL\s*[.:/]\s*([\d.,]+)",
    r"GRAND\s+TOTAL\s*[.:/]?\s*([\d.,]+)",
    r"TOTAL\s+INVOICE\s*[.:/]?\s*([\d.,]+)",
]

# Destinatario
PATRONES_DESTINATARIO = [
    r"SE[ÑN]OR\(ES\)\s*/\s*MESSRS\s*:\s*(.+?)\s+FECHA",   # formato AquaChile/Blumar
    r"(?:CONSIGNEE|DESTINATARIO|BUYER|COMPRADOR)\s*[.:/]\s*(.+?)(?:\n\n|\n[A-Z]{2,}:)",
    r"SOLD\s+TO\s*[.:/]\s*(.+?)(?:\n\n|\n[A-Z])",
    r"SHIP\s+TO\s*[.:/]\s*(.+?)(?:\n\n|\n[A-Z])",
]

# Dirección del destinatario
PATRONES_DIRECCION = [
    r"DIRECCI[OÓ]N\s*/\s*ADDRESS\s*:\s*(.*?)\s*(?:GIRO\s*:|CIUDAD\s*/\s*CITY\s*:)",
    r"ADDRESS\s*[.:/]\s*(.+?)(?:\n\n|\n[A-Z]{2,}:)",
]

# País de destino — para generar glosa casilla 18
PATRONES_PAIS = [
    r"COUNTRY\s+OF\s+(?:FINAL\s+)?DESTINATION\s*[.:/]\s*([A-Z\s]+?)(?:\n|,)",
    r"DESTINO\s+FINAL\s*[.:/]\s*([A-Z\s]+?)(?:\n|,)",
    r"\b(USA|CHINA|VIETNAM|MEXICO|M[EÉ]XICO|TAIWAN|RUSSIA|COLOMBIA|PANAMA|ARGENTINA)\b",
]

# Ciudad/puerto de destino — para separar CRTs cuando mismo cliente va a distintas ciudades
PATRONES_CIUDAD_DESTINO = [
    r"PORT\s+OF\s+(?:DISCHARGE|DESTINATION|ENTRY|DELIVERY)\s*[.:/]\s*([A-Z][A-Z\s,]+?)(?:\n|,|\s{2,}|$)",
    r"DELIVERY\s+PORT\s*[.:/]\s*([A-Z][A-Z\s]+?)(?:\n|,)",
    r"PLACE\s+OF\s+DELIVERY\s*[.:/]\s*([A-Z][A-Z\s]+?)(?:\n|,)",
]

_CIUDADES_CONOCIDAS = [
    "MIAMI", "LOS ANGELES", "NEW YORK", "SEATTLE", "BOSTON",
    "CHICAGO", "HOUSTON", "SAN FRANCISCO", "ATLANTA",
]


def _extraer_ciudad(texto_up: str, direccion: Optional[str]) -> str:
    """
    Extrae la ciudad/puerto de destino.
    Primero busca patrones explícitos (PORT OF DISCHARGE etc.),
    luego busca ciudades conocidas en la dirección del consignatario.
    Retorna "" si no puede determinarlo.
    """
    raw = _primera_coincidencia(texto_up, PATRONES_CIUDAD_DESTINO)
    if raw:
        return raw.strip().rstrip(",. ").upper()
    if direccion:
        dir_up = direccion.upper()
        for ciudad in _CIUDADES_CONOCIDAS:
            if ciudad in dir_up:
                return ciudad
    return ""


# Certificado sanitario en facturas (algunas lo incluyen)
PATRONES_CERT = [
    r"CERTIFICADO\s+SANITARIO\s*(?:NRO?|N[°º])\s*[.:/]?\s*(\d+)",
    r"N[°º]\s+CODAUT\s*[.:/]\s*(\d+)",
    r"CODAUT\s*[.:/]\s*(\d+)",
]

# Referencia cruzada con la guía — número que coincide con orden_venta en la guía
# Por pesquera, según los documentos reales
PATRONES_REF_PEDIDO_PESQUERA = {
    # Multi X: factura dice "PV: 12345" — guía dice "N° Pedido: 12345"
    "multix":            [r"PV\s*[.:]\s*(\d+)"],
    # AquaChile: factura dice "N° PEDIDO / ORDER: 12345"
    "aquachile":         [r"N[°º]\s*PEDIDO\s*/\s*ORDER\s*[.:/]?\s*(\d+)"],
    # Blumar: valor debajo de "SELLER'S REFERENCE No."
    "blumar":            [r"SELLER'S\s+REFERENCE\s+N[Oo°]\.?\s*[:\n]?\s*(\d+)",
                          r"SELLER.S\s+REFERENCE\s+N[Oo°]\.?\s*[:\n]?\s*(\d+)"],
    "blumar_magallanes": [r"SELLER'S\s+REFERENCE\s+N[Oo°]\.?\s*[:\n]?\s*(\d+)",
                          r"SELLER.S\s+REFERENCE\s+N[Oo°]\.?\s*[:\n]?\s*(\d+)"],
    # Cermaq: factura dice "Order Ref: 12345"
    "cermaq":            [r"ORDER\s+REF\.?\s*[.:/]?\s*(\d+)",
                          r"Order\s+Ref\.?\s*[.:/]?\s*(\d+)"],
    # Australis: fallback
    "australis":         [r"N[°º]\s*PEDIDO\s*[.:/]\s*(\d+)",
                          r"ORDER\s+N[°º]?\s*[.:/]\s*(\d+)"],
}

INCOTERMS_VALIDOS = {"CPT", "CFR", "CIF", "FOB", "FCA", "DAP", "EXW", "DDP"}

# Fecha de emisión del documento
PATRONES_FECHA = [
    r"FECHA\s+(?:DE\s+)?(?:EMISION|EMISI[OÓ]N|DOCUMENTO|FACTURA)\s*[.:/]\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    r"DATE\s*[.:/]\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    r"INVOICE\s+DATE\s*[.:/]\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    r"(\d{1,2}\s+(?:DE\s+)?(?:ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(?:DE\s+)?\d{4})",
    r"(\d{1,2}\s+(?:JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+\d{4})",
    # dd-mm-yyyy o dd/mm/yyyy genérico — último recurso
    r"(\d{2}[/-]\d{2}[/-]\d{4})",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _limpiar_numeros(texto: str) -> str:
    nums = re.findall(r'\d+', texto)
    return ", ".join(nums) if nums else ""


def _primera_coincidencia(texto: str, patrones: list) -> Optional[str]:
    for patron in patrones:
        m = re.search(patron, texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
        if m:
            return m.group(1).strip()
    return None


def _extraer_float(texto: str) -> Optional[float]:
    if not texto:
        return None
    t = texto.replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        pass
    try:
        return float(texto.replace(",", ""))
    except ValueError:
        return None


_MESES_ES = {
    "ENERO": "01", "FEBRERO": "02", "MARZO": "03", "ABRIL": "04",
    "MAYO": "05", "JUNIO": "06", "JULIO": "07", "AGOSTO": "08",
    "SEPTIEMBRE": "09", "OCTUBRE": "10", "NOVIEMBRE": "11", "DICIEMBRE": "12",
}
_MESES_EN = {
    "JANUARY": "01", "FEBRUARY": "02", "MARCH": "03", "APRIL": "04",
    "MAY": "05", "JUNE": "06", "JULY": "07", "AUGUST": "08",
    "SEPTEMBER": "09", "OCTOBER": "10", "NOVEMBER": "11", "DECEMBER": "12",
}


def _normalizar_fecha(raw: str) -> Optional[str]:
    """
    Convierte cualquier formato de fecha a 'DD-MM-YYYY'.
    Soporta: dd/mm/yyyy, dd-mm-yyyy, '12 de Marzo de 2025', '12 March 2025'.
    Retorna None si no puede interpretar.
    """
    if not raw:
        return None
    raw = raw.strip().upper()

    # Texto con nombre de mes (español o inglés)
    meses = {**_MESES_ES, **_MESES_EN}
    for nombre, numero in meses.items():
        if nombre in raw:
            nums = re.findall(r'\d+', raw)
            if len(nums) >= 2:
                dia = nums[0].zfill(2)
                anio = nums[-1]
                if len(anio) == 2:
                    anio = "20" + anio
                return f"{dia}-{numero}-{anio}"

    # Numérico: dd/mm/yyyy o dd-mm-yyyy
    m = re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', raw)
    if m:
        dia, mes, anio = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        if len(anio) == 2:
            anio = "20" + anio
        return f"{dia}-{mes}-{anio}"

    return None


def _normalizar_pais(raw: str) -> str:
    """Normaliza el país para la glosa de casilla 18."""
    mapa = {
        "USA": "USA", "ESTADOS UNIDOS": "USA", "UNITED STATES": "USA",
        "CHINA": "CHINA", "VIETNAM": "VIETNAM", "VIET NAM": "VIETNAM",
        "MEXICO": "MEXICO", "MÉXICO": "MEXICO",
        "TAIWAN": "TAIWAN", "RUSSIA": "RUSSIA", "RUSIA": "RUSSIA",
        "COLOMBIA": "COLOMBIA", "PANAMA": "PANAMA", "PANAMÁ": "PANAMA",
        "ARGENTINA": "ARGENTINA",
    }
    raw_up = raw.strip().upper()
    for k, v in mapa.items():
        if k in raw_up:
            return v
    return raw_up


# ── Extractor de productos para facturas Australis ───────────────────────────
#
# Las facturas Australis tienen las líneas de producto en inglés.
# Ejemplo:  "17 CAJAS  FRESH ATLANTIC SALMON TRIM D FILLET PREMIUM 2-3 LBS 35 LB  286.05 KG"
#
# pdfplumber suele extraer la tabla con columnas:
#   QTY/BOXES  |  DESCRIPTION  |  NET KG  |  ...
# Si no puede extraer la tabla, el fallback regex busca líneas que contengan
# "SALMON" + número de cajas + kilos.


def _extraer_productos_australis(texto: str, pdf_path: str) -> list[dict]:
    """
    Extrae las líneas de producto de una factura Australis (en inglés).

    Retorna lista de dicts:
        descripcion    — texto inglés del producto (ej: "FRESH ATLANTIC SALMON TRIM D FILLET PREMIUM 2-3 LBS 35 LB")
        familia        — igual a descripcion para Australis (se usa completo en casilla 11)
        cajas_totales  — int
        kilos_totales  — float (kg netos)
    """
    # ── Intento 1: tabla pdfplumber ───────────────────────────────────────────
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    header = [str(c or "").upper().strip() for c in table[0]]
                    # Tabla Australis: DESCRIPTION/PRODUCT + QTY/BOXES + NET KG/NET WEIGHT
                    has_desc = any("DESCRIPTION" in h or "PRODUCT" in h for h in header)
                    has_qty  = any("QTY" in h or "BOX" in h or "CASES" in h or "CARTON" in h for h in header)
                    if not has_desc or not has_qty:
                        continue

                    idx_desc  = next(i for i, h in enumerate(header) if "DESCRIPTION" in h or "PRODUCT" in h)
                    idx_qty   = next(i for i, h in enumerate(header) if "QTY" in h or "BOX" in h or "CASES" in h or "CARTON" in h)
                    idx_kg    = next(
                        (i for i, h in enumerate(header) if "NET" in h and ("KG" in h or "WEIGHT" in h or "KILO" in h)),
                        next((i for i, h in enumerate(header) if "KG" in h or "KILO" in h), None),
                    )

                    productos = []
                    for row in table[1:]:
                        if not row:
                            continue
                        desc_raw = str(row[idx_desc] or "").strip()
                        qty_raw  = str(row[idx_qty] or "").strip()

                        if not desc_raw or "TOTAL" in desc_raw.upper():
                            continue
                        if not any(c.isalpha() for c in desc_raw):
                            continue
                        if "SALMON" not in desc_raw.upper():
                            continue

                        try:
                            cajas = int(float(qty_raw.replace(",", "").replace(".", "")))
                        except (ValueError, AttributeError):
                            cajas = 0
                        if cajas <= 0:
                            continue

                        kilos = None
                        if idx_kg is not None and idx_kg < len(row):
                            kilos = _extraer_float(str(row[idx_kg] or ""))

                        desc = re.sub(r'\s+', ' ', desc_raw).strip().upper()
                        productos.append({
                            "descripcion":   desc,
                            "familia":       desc,
                            "kilos_totales": kilos or 0.0,
                            "cajas_totales": cajas,
                        })

                    if productos:
                        return productos

    except Exception as e:
        print(f"[extractor_facturas] productos Australis tabla error: {e}")

    # ── Intento 2: regex sobre texto ──────────────────────────────────────────
    productos = []
    # Patrón: número + descripción con SALMON + kilos
    patron = re.compile(
        r'(\d+)\s+'                              # cajas
        r'((?:FRESH\s+)?ATLANTIC\s+SALMON[^\n]{5,100}?'  # descripción
        r'(?:\d+[-–]\d+\s*(?:LBS?|KGS?)[^\n]{0,40})?)'   # con talla
        r'\s+([\d.,]+)\s*(?:KG|KILOS?)?',        # kilos netos
        re.IGNORECASE
    )
    for m in patron.finditer(texto):
        try:
            cajas = int(m.group(1))
        except ValueError:
            cajas = 0
        desc  = re.sub(r'\s+', ' ', m.group(2)).strip().upper()
        kilos = _extraer_float(m.group(3))
        if cajas > 0 and desc:
            productos.append({
                "descripcion":   desc,
                "familia":       desc,
                "kilos_totales": kilos or 0.0,
                "cajas_totales": cajas,
            })

    return productos


# ── Extractor Excel (Blumar y Cermaq envían facturas en .xlsx) ────────────────

def _buscar_celda_excel(ws, texto: str):
    """Busca celda cuyo valor contenga 'texto'. Retorna (row, col) o None."""
    texto_up = texto.upper()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and texto_up in str(cell.value).upper():
                return cell.row, cell.column
    return None


def _celda_excel(ws, row, col) -> Optional[str]:
    """Retorna valor de celda como string limpio o None."""
    try:
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None and str(v).strip() not in ("", "None") else None
    except Exception:
        return None


def _texto_excel(ws) -> str:
    """Extrae todo el texto de la hoja como string para detección de pesquera/país."""
    partes = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                partes.append(str(cell.value))
    return "\n".join(partes)


def extraer_datos_factura_excel(path: str) -> Optional[dict]:
    """
    Extrae datos de una factura de exportación en formato Excel (.xlsx).
    Soporta los formatos de Blumar y Cermaq.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        texto_completo = _texto_excel(ws)
        texto_up = texto_completo.upper()

        from modulos.config_cliente import detectar_pesquera
        pesquera = detectar_pesquera(texto_up)

        # ── Número de factura ─────────────────────────────────────────────────
        numero_factura = None
        for label in ["INVOICE NO", "INVOICE N°", "N° FACTURA", "FACTURA N°",
                       "NUMERO FACTURA", "COMMERCIAL INVOICE"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                # Probar celda derecha (col+1) y debajo (row+1)
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v and re.search(r'\d', v):
                    numero_factura = re.sub(r'\D', '', v.split()[0]) or v
                    break

        # ── Referencia cruzada con guía (ref_pedido) ──────────────────────────
        ref_pedido = None
        labels_ref = {
            "blumar":            ["SELLER'S REFERENCE", "SELLERS REFERENCE"],
            "blumar_magallanes": ["SELLER'S REFERENCE", "SELLERS REFERENCE"],
            "cermaq":            ["ORDER REF", "ORDER REFERENCE"],
        }
        for label in labels_ref.get(pesquera, []):
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    m = re.search(r'(\d+)', v)
                    if m:
                        ref_pedido = m.group(1)
                        break

        # ── Incoterm ──────────────────────────────────────────────────────────
        incoterm = None
        for label in ["INCOTERM", "CLAUSULA DE VENTA", "DELIVERY TERM"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    m = re.search(r'\b(CPT|CFR|CIF|FOB|FCA|DAP|EXW|DDP)\b', v.upper())
                    if m:
                        incoterm = m.group(1)
                        break
        if not incoterm:
            m = re.search(r'\b(CPT|CFR|CIF|FOB|FCA|DAP|EXW|DDP)\b', texto_up)
            if m:
                incoterm = m.group(1)

        # ── Total ─────────────────────────────────────────────────────────────
        total = None
        for label in ["GRAND TOTAL", "TOTAL INVOICE", "TOTAL AMOUNT", "TOTAL USD",
                       "TOTAL GENERAL", "MONTO TOTAL"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    total = _extraer_float(v)
                    if total:
                        break

        # ── Destinatario ──────────────────────────────────────────────────────
        destinatario = None
        for label in ["CONSIGNEE", "BUYER", "SOLD TO", "DESTINATARIO", "SHIP TO"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v and len(v) > 3:
                    destinatario = v
                    break

        # ── Moneda ────────────────────────────────────────────────────────────
        moneda = None
        m = re.search(r'\b(USD|CNY|EUR|CLP)\b', texto_up)
        if m:
            moneda = m.group(1)

        # ── País destino ──────────────────────────────────────────────────────
        raw_pais = None
        for label in ["COUNTRY OF DESTINATION", "FINAL DESTINATION", "DESTINO FINAL"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    raw_pais = v
                    break
        if not raw_pais:
            m = re.search(r'\b(USA|CHINA|VIETNAM|MEXICO|TAIWAN|RUSSIA|COLOMBIA|PANAMA)\b', texto_up)
            if m:
                raw_pais = m.group(1)
        pais_destino = _normalizar_pais(raw_pais) if raw_pais else "USA"

        # ── Ciudad/puerto de destino ──────────────────────────────────────────
        ciudad_destino = _extraer_ciudad(texto_up, None)

        # ── Bultos (para detección de discrepancias) ──────────────────────────
        bultos = None
        for label in ["TOTAL CAJAS", "TOTAL BOXES", "BULTOS", "PACKAGES", "UNITS"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    try:
                        bultos = int(float(v.replace(",", "").replace(".", "")))
                        break
                    except Exception:
                        pass

        # ── Fecha ─────────────────────────────────────────────────────────────
        fecha = None
        for label in ["INVOICE DATE", "FECHA", "DATE", "FECHA EMISION"]:
            pos = _buscar_celda_excel(ws, label)
            if pos:
                v = _celda_excel(ws, pos[0], pos[1] + 1) or _celda_excel(ws, pos[0] + 1, pos[1])
                if v:
                    fecha = _normalizar_fecha(v)
                    if fecha:
                        break
        if not fecha:
            raw_f = _primera_coincidencia(texto_up, PATRONES_FECHA)
            fecha = _normalizar_fecha(raw_f) if raw_f else None

        # ── Productos (solo Australis — resto retorna lista vacía) ──────────────
        productos = []
        if pesquera == "australis":
            productos = _extraer_productos_australis(texto_completo, path)

        return {
            "tipo":           "factura",
            "pesquera":       pesquera,
            "numero_factura": numero_factura,
            "ref_pedido":     ref_pedido,
            "incoterm":       incoterm,
            "moneda":         moneda,
            "total":          total,
            "destinatario":   destinatario,
            "direccion":      None,
            "pais_destino":   pais_destino,
            "ciudad_destino": ciudad_destino,
            "cert_sanitario": None,
            "bultos":         bultos,
            "fecha":          fecha,
            "productos":      productos,
            "texto_completo": texto_completo,
        }

    except Exception as e:
        print(f"[extractor_facturas_excel] Error: {e}")
        return None


# ── Extractor PDF principal ───────────────────────────────────────────────────

def extraer_datos_factura(path: str) -> Optional[dict]:
    """
    Extrae datos de una factura de exportación.
    Soporta PDF y Excel (.xlsx).
    """
    if path.lower().endswith((".xlsx", ".xls")):
        return extraer_datos_factura_excel(path)

    try:
        texto = ""
        with pdfplumber.open(path) as pdf:
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages)

        if not texto.strip():
            return None

        texto_up = texto.upper()

        # Detectar pesquera primero — necesario para ref_pedido
        from modulos.config_cliente import detectar_pesquera
        pesquera = detectar_pesquera(texto_up)

        # Número de factura
        raw_fact = _primera_coincidencia(texto_up, PATRONES_FACTURA)
        if not raw_fact:
            m_blumar = re.search(PATRON_FACTURA_BLUMAR, texto_up)
            if m_blumar:
                raw_fact = m_blumar.group(1)
        numero_factura = _limpiar_numeros(raw_fact) if raw_fact else None

        # Referencia cruzada con la guía (ref_pedido) — patrón específico por pesquera
        patrones_ref = PATRONES_REF_PEDIDO_PESQUERA.get(pesquera, [])
        raw_ref    = _primera_coincidencia(texto_up, patrones_ref) if patrones_ref else None
        ref_pedido = raw_ref.strip() if raw_ref else None

        # Incoterm
        raw_inco = _primera_coincidencia(texto_up, PATRONES_INCOTERM)
        incoterm = raw_inco.strip().upper() if raw_inco else None
        if incoterm and incoterm not in INCOTERMS_VALIDOS:
            incoterm = None

        # Moneda
        raw_moneda = _primera_coincidencia(texto_up, PATRONES_MONEDA)
        moneda = raw_moneda.strip().upper() if raw_moneda else None

        # Total
        raw_total = _primera_coincidencia(texto_up, PATRONES_TOTAL)
        total = _extraer_float(raw_total)

        # Destinatario
        raw_dest     = _primera_coincidencia(texto, PATRONES_DESTINATARIO)
        destinatario = raw_dest.strip() if raw_dest else None

        # Dirección
        raw_dir = _primera_coincidencia(texto, PATRONES_DIRECCION)
        if raw_dir:
            raw_dir = re.sub(r"TRANSPORTE\s*/\s*TRANSPORT\s*:\s*\S+", "", raw_dir, flags=re.IGNORECASE)
            direccion = re.sub(r"\s+", " ", raw_dir).strip()
        else:
            direccion = None

        # País destino
        raw_pais     = _primera_coincidencia(texto_up, PATRONES_PAIS)
        pais_destino = _normalizar_pais(raw_pais) if raw_pais else "USA"

        # Ciudad/puerto de destino
        ciudad_destino = _extraer_ciudad(texto_up, direccion)

        # Certificado (algunas facturas lo incluyen)
        raw_cert       = _primera_coincidencia(texto_up, PATRONES_CERT)
        cert_sanitario = raw_cert.strip() if raw_cert else None

        # Fecha del documento
        raw_fecha = _primera_coincidencia(texto_up, PATRONES_FECHA)
        fecha     = _normalizar_fecha(raw_fecha) if raw_fecha else None

        # Productos (solo Australis — resto retorna lista vacía)
        productos = []
        if pesquera == "australis":
            productos = _extraer_productos_australis(texto, path)

        return {
            "tipo":           "factura",
            "pesquera":       pesquera,
            "numero_factura": numero_factura,
            "ref_pedido":     ref_pedido,
            "incoterm":       incoterm,
            "moneda":         moneda,
            "total":          total,
            "destinatario":   destinatario,
            "direccion":      direccion,
            "pais_destino":    pais_destino,
            "ciudad_destino":  ciudad_destino,
            "cert_sanitario":  cert_sanitario,
            "fecha":           fecha,
            "productos":       productos,
            "texto_completo":  texto,
        }

    except Exception as e:
        print(f"[extractor_facturas] Error: {e}")
        return None
