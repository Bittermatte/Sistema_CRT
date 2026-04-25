"""
Corrige los problemas visuales del Excel plantilla CRT.
Ejecutar una sola vez: python3 scripts/fix_plantilla.py
"""
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

PLANTILLA = Path("plantillas/crt_plantilla.xlsx")

wb = load_workbook(PLANTILLA)
ws = wb.active

# ── Fuentes base ──────────────────────────────────────────────────────────────
FONT_LABEL    = Font(name="Arial",    size=6,    bold=False)   # labels casillas
FONT_DATA     = Font(name="Calibri",  size=9,    bold=False)   # datos variables principales
FONT_DATA_SM  = Font(name="Calibri",  size=8,    bold=False)   # datos secundarios
FONT_DATA_XS  = Font(name="Calibri",  size=7,    bold=False)   # descripción carga, docs
FONT_DATA_BD  = Font(name="Calibri",  size=9,    bold=True)    # totales, patentes
FONT_NUM_CRT  = Font(name="Calibri",  size=12,   bold=True)    # número CRT grande
FONT_TRANSP   = Font(name="Calibri",  size=11,   bold=True,    italic=True)  # transportista
FONT_FLETE_LB = Font(name="Arial",    size=9,    bold=False)   # "Flete / Frete" label
FONT_HEADER   = Font(name="Arial",    size=18,   bold=True)    # "CRT" título
FONT_FIRMA    = Font(name="Arial",    size=10,   bold=True)    # firma transportista pie

ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_WRAP    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

# ── 1. Eliminar la "A" sobre el recuadro ─────────────────────────────────────
# La celda A1 tiene "A" (marca ALADI) que aparece fuera del recuadro en el PDF
ws["A1"] = None

# ── 2. Fuente y alineación del número CRT (F13) ───────────────────────────────
ws["F13"].font      = FONT_NUM_CRT
ws["F13"].alignment = ALIGN_CENTER

# ── 3. Fuente transportista (F16) — BoldItalic ───────────────────────────────
ws["F16"].font      = FONT_TRANSP
ws["F16"].alignment = ALIGN_CENTER
ws["F17"].font      = FONT_TRANSP
ws["F17"].alignment = ALIGN_CENTER
ws["F18"].font      = FONT_TRANSP
ws["F18"].alignment = ALIGN_CENTER

# ── 4. Datos variables — fuente Calibri 9 uniforme ───────────────────────────
data_cells_main = [
    "A13", "A14", "A15",           # remitente
    "A18", "A19", "A20", "A21",    # destinatario
    "A23", "A24", "A25", "A26",    # consignatario
    "A28", "A29", "A30",           # notificar
    "F21",                          # lugar emisión
    "F24",                          # lugar recepción + fecha
    "F27",                          # lugar entrega
    "F28",                          # destino final
]
for coord in data_cells_main:
    ws[coord].font      = FONT_DATA
    ws[coord].alignment = ALIGN_CENTER

# ── 5. Wrap text en direcciones (campos 4, 6, 9) ─────────────────────────────
wrap_cells = ["A18","A19","A20","A21","A23","A24","A25","A26","A28","A29","A30"]
for coord in wrap_cells:
    ws[coord].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws[coord].font = FONT_DATA_SM

# ── 6. "SON:" y productos — indent 2 (reducido a la mitad) ─────────────────
# A34 = "SON:" label, A35/A37 = descripción, A36/A38 = kilos
ws["A34"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws["A34"].font      = Font(name="Calibri", size=8, bold=False)

for coord in ["A35", "A36", "A37", "A38"]:
    ws[coord].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws[coord].font      = Font(name="Calibri", size=8, bold=False)

# ── 7. "Flete / Frete" — misma fuente y tamaño que ORIGEN/FRONTERA ───────────
# A46 = "Flete / Frete", A47 = "ORIGEN/FRONTERA", A48 = "FRONTERA/DESTINO"
ws["A46"].font      = Font(name="Arial", size=6, bold=False)
ws["A46"].alignment = ALIGN_LEFT
ws["A47"].font      = Font(name="Arial", size=6, bold=False)
ws["A48"].font      = Font(name="Arial", size=6, bold=False)

# Valores de flete B47/B48/B52 — Calibri mismo tamaño
for coord in ["B47", "B48", "B52"]:
    ws[coord].font      = Font(name="Calibri", size=9, bold=False)
    ws[coord].alignment = ALIGN_CENTER
ws["C47"].font = Font(name="Calibri", size=9, bold=False)
ws["C48"].font = Font(name="Calibri", size=9, bold=False)
ws["C52"].font = Font(name="Calibri", size=9, bold=False)

# ── 8. TOTAL en sección 15 — alineación coherente ────────────────────────────
# A52 = "T O T A L" — mismo estilo que A47/A48
ws["A52"].font      = Font(name="Arial", size=9, bold=False)
ws["A52"].alignment = ALIGN_LEFT

# ── 9. Totales casilla 14 (A40/A41/A42) — alineación izquierda uniforme ──────
for coord in ["A40", "A41", "A42"]:
    ws[coord].font      = Font(name="Calibri", size=8, bold=True)
    ws[coord].alignment = ALIGN_LEFT

# G40 (valor + incoterm) — alineación derecha
ws["G40"].font      = Font(name="Calibri", size=9, bold=False)
ws["G40"].alignment = ALIGN_CENTER

# ── 10. Firma transportista pie (A71) ────────────────────────────────────────
ws["A71"].font      = FONT_FIRMA
ws["A71"].alignment = ALIGN_CENTER

# ── 11. Área de impresión — asegurarse que cubre solo el formulario ───────────
# Excluir filas superiores si la "A" estaba en fila 1
ws.print_area = "A2:L75"

# ── 12. Guardar ──────────────────────────────────────────────────────────────
wb.save(PLANTILLA)
print(f"Plantilla corregida: {PLANTILLA}")
