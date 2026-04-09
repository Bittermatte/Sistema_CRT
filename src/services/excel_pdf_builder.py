"""
Excel PDF Builder — genera CRT llenando la plantilla Excel y convirtiendo a PDF.
Estrategia: openpyxl llena celdas → LibreOffice headless convierte a PDF.
El PDF resultante es idéntico al original de Google Sheets.
"""

import io
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins

PLANTILLA_PATH = Path("plantillas/crt_plantilla.xlsx")

# Rutas de LibreOffice según plataforma
_SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS brew cask
    "/usr/local/bin/soffice",                                  # macOS symlink
    "/opt/homebrew/bin/soffice",                               # macOS M1/M2
    "/usr/bin/libreoffice",                                    # Linux
    "/usr/bin/soffice",                                        # Linux alternativo
]

def _find_soffice() -> Optional[str]:
    for candidate in _SOFFICE_CANDIDATES:
        if shutil.which(candidate) or Path(candidate).exists():
            return candidate
    return None


def _fmt(value) -> str:
    """Formatea un valor del formulario a string limpio."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, float):
        return "" if value == 0.0 else f"{value:,.2f}"
    if isinstance(value, int):
        return "" if value == 0 else str(value)
    text = str(value).strip()
    if " — " in text:
        text = text.split(" — ")[0]
    return text


def _get(form: dict, key: str) -> str:
    return _fmt(form.get(key))


def _kn_cell(kn: str) -> str:
    """
    Formatea el valor de kilos netos para la celda del Excel.
    Si ya viene con prefijo 'CON:' (AquaChile), lo retorna tal cual.
    Si viene vacío (Blumar/Cermaq/Australis/Multi X — kilos inline), retorna "".
    """
    if not kn:
        return ""
    return kn if kn.upper().startswith("CON:") else f"CON: {kn} KILOS NETOS"


def _fill_workbook(form: dict) -> openpyxl.Workbook:
    """
    Carga la plantilla Excel y llena las celdas variables con los datos del formulario.
    Devuelve el workbook modificado sin tocar el archivo original.
    """
    wb = load_workbook(PLANTILLA_PATH)
    ws = wb.active

    # ── Casilla 2 — Número CRT ────────────────────────────────────────────────
    ws["F13"] = _get(form, "f_numero_crt")

    # ── Casilla 1 — Remitente ─────────────────────────────────────────────────
    ws["A13"] = _get(form, "f_remitente")
    dir_rem = _get(form, "f_dir_remitente")
    if "\n" in dir_rem:
        parts = dir_rem.split("\n", 1)
        ws["A14"] = parts[0].strip()
        ws["A15"] = parts[1].strip()
    else:
        ws["A14"] = dir_rem

    # ── Casilla 3 — Transportista (desde form_data, no hardcodeado) ───────────
    ws["F16"] = _get(form, "f_transportista")
    dir_trans  = _get(form, "f_dir_transportista") or "Avda Colon 1761 Bahia Blanca BS - AS\nARGENTINA"
    lines_trans = [l.strip() for l in dir_trans.split("\n") if l.strip()]
    ws["F17"] = lines_trans[0] if len(lines_trans) > 0 else ""
    ws["F18"] = lines_trans[1] if len(lines_trans) > 1 else ""

    # ── Casilla 4 — Destinatario ──────────────────────────────────────────────
    ws["A18"] = _get(form, "f_destinatario")
    dir_dest = _get(form, "f_dir_destinatario")
    if dir_dest:
        lines = [l.strip() for l in dir_dest.split("\n") if l.strip()]
        # A19:E20 es merge de 2 filas — combinar líneas 1+2 con \n en A19
        ws["A19"] = "\n".join(lines[:2]) if len(lines) >= 2 else (lines[0] if lines else "")
        # A21:E21 para la tercera línea si existe
        if len(lines) > 2: ws["A21"] = lines[2]

    # ── Casilla 5 — Lugar Emisión ─────────────────────────────────────────────
    ws["F21"] = _get(form, "f_lugar_emision")

    # ── Casilla 6 — Consignatario ─────────────────────────────────────────────
    ws["A23"] = _get(form, "f_consignatario")
    dir_cons = _get(form, "f_dir_consignatario")
    if dir_cons:
        lines = [l.strip() for l in dir_cons.split("\n") if l.strip()]
        if len(lines) > 0: ws["A24"] = lines[0]
        if len(lines) > 1: ws["A25"] = lines[1]
        if len(lines) > 2: ws["A26"] = lines[2]

    # ── Casilla 7 — Lugar y fecha recepción ──────────────────────────────────
    lugar_rec = _get(form, "f_lugar_recepcion")
    fecha_doc = _get(form, "f_fecha_documento")
    ws["F24"] = f"{lugar_rec}  {fecha_doc}".strip() if lugar_rec or fecha_doc else ""

    # ── Casilla 8 — Lugar entrega ─────────────────────────────────────────────
    ws["F27"] = _get(form, "f_lugar_entrega")

    # ── Casilla 9 — Notificar / Destino final ────────────────────────────────
    ws["A28"] = _get(form, "f_notificar")
    ws["F28"] = _get(form, "f_destino_final")
    dir_not = _get(form, "f_dir_notificar")
    if dir_not:
        lines = [l.strip() for l in dir_not.split("\n") if l.strip()]
        # A29:E30 es merge de 2 filas — combinar con \n en A29
        ws["A29"] = "\n".join(lines[:2]) if len(lines) >= 2 else (lines[0] if lines else "")

    # ── Casilla 11 — Descripción de carga (hasta 5 líneas) ───────────────────
    # Filas A35..A39: contenido de producto (desc + kn opcionales)
    # Filas A40-A42: TOTALES (posición fija en el template)
    #
    # Pesqueras con kn inline (Blumar/Cermaq/Australis/Multi X): 1 fila por producto
    # → caben hasta 5 productos en las filas 35-39
    # AquaChile con kn separado: 2 filas por producto → caben hasta 2 productos en 35-38
    row_y = 35
    for n in range(1, 6):
        desc_n = _get(form, f"f_descripcion_{n}")
        kn_n   = _kn_cell(_get(form, f"f_kilos_netos_{n}"))
        if not desc_n:
            break
        if row_y > 39:    # no sobrepasar los TOTALES
            break
        ws.cell(row=row_y, column=1).value = desc_n
        row_y += 1
        if kn_n and row_y <= 39:
            ws.cell(row=row_y, column=1).value = kn_n
            row_y += 1

    # ── Casilla 12 — Peso bruto ───────────────────────────────────────────────
    pb = _get(form, "f_peso_bruto")
    ws["G34"] = f"{pb} Kg Brutos" if pb else "Kg Brutos"

    # ── Totales (posición fija) ───────────────────────────────────────────────
    total_cajas = _get(form, "f_total_cajas")
    total_kn    = _get(form, "f_peso_neto")
    total_kb    = _get(form, "f_peso_bruto")
    ws["A40"] = f"     TOTAL CAJAS:  {total_cajas}"
    ws["A41"] = f"     TOTAL KILOS NETOS: {total_kn}"
    ws["A42"] = f"     TOTAL KILOS BRUTOS: {total_kb}"

    # ── Casilla 14 — Valor / Incoterm ─────────────────────────────────────────
    valor = _get(form, "f_valor_mercaderia")
    inco  = _get(form, "f_incoterm")
    ws["G40"] = f"  {valor}  {inco}".strip() if valor else f"  {inco}"

    # ── Casilla 16 — Declaración valor ───────────────────────────────────────
    ws["F45"] = f"US$  {valor}" if valor else "US$"

    # ── Casilla 15 — Fletes desglosados + total ───────────────────────────────
    flete_orig = _get(form, "f_flete_origen")
    flete_fron = _get(form, "f_flete_frontera")
    flete_tot  = _get(form, "f_flete_usd")
    if flete_orig:
        try:
            ws["B47"] = float(flete_orig.replace(".", "").replace(",", "."))
        except (ValueError, AttributeError):
            ws["B47"] = flete_orig
    if flete_fron:
        try:
            ws["B48"] = float(flete_fron.replace(".", "").replace(",", "."))
        except (ValueError, AttributeError):
            ws["B48"] = flete_fron
    if flete_tot:
        try:
            ws["B52"] = float(flete_tot.replace(".", "").replace(",", "."))
        except (ValueError, AttributeError):
            ws["B52"] = flete_tot

    # ── Casilla 17 — Documentos anexos ───────────────────────────────────────
    facturas = _get(form, "f_num_factura")
    guias    = _get(form, "f_guias_despacho")
    cert_san = _get(form, "f_cert_sanitario")
    ws["F49"] = f"FACTURAS NROS:  {facturas}" if facturas else "FACTURAS NROS:"
    ws["F50"] = f"GUIAS DE DESPACHO NROS:  {guias}" if guias else "GUIAS DE DESPACHO NROS:"
    ws["F51"] = f"CERTIFICADO SANITARIO NRO:  {cert_san}" if cert_san else "CERTIFICADO SANITARIO NRO:"

    # ── Casilla 18 — Instrucciones aduana ────────────────────────────────────
    instrucciones = _get(form, "f_instrucciones_aduana")
    if instrucciones:
        lines = instrucciones.split("\n")
        ws["F55"] = lines[0].strip() if len(lines) > 0 else ""
        ws["F56"] = lines[1].strip() if len(lines) > 1 else ""
        ws["F57"] = lines[2].strip() if len(lines) > 2 else ""

    # ── Casilla 21 — Firma remitente por pesquera ─────────────────────────────
    firma_rem = _get(form, "f_firma_remitente")
    ws["A60"] = firma_rem if firma_rem else _get(form, "f_remitente")

    # ── Casilla 21/23 — Fechas firma ─────────────────────────────────────────
    fecha_em = _get(form, "f_fecha_emision")
    ws["A62"] = f"Fecha / Data     {fecha_em}" if fecha_em else "Fecha / Data"

    # ── Casilla 22 — Conductor y patentes ────────────────────────────────────
    conductor  = _get(form, "f_conductor")
    pat_camion = _get(form, "f_patente_camion")
    pat_rampla = _get(form, "f_patente_rampla")
    ws["F62"] = f"CONDUCTOR:  {conductor}" if conductor else "CONDUCTOR:"
    ws["F63"] = f"PATENTE CAMION:  {pat_camion}  / PATENTE RAMPLA:  {pat_rampla}"

    # ── Página — sin márgenes, una sola página A4 ─────────────────────────────
    ws.page_margins = PageMargins(
        left=0.0, right=0.0, top=0.0, bottom=0.0, header=0.0, footer=0.0
    )
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # fitToPage vía pageSetUpPr (openpyxl bug workaround)
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize   = 9   # A4
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A2:L75"

    # ── Altura de filas para direcciones multilínea ───────────────────────────
    for row_num in [18, 19, 20, 21, 23, 24, 25, 26, 28, 29, 30]:
        ws.row_dimensions[row_num].height = 14

    return wb


def generate_crt_pdf_from_excel(form_data: dict) -> Optional[bytes]:
    """
    Genera el PDF del CRT llenando la plantilla Excel y convirtiendo con LibreOffice.
    Devuelve bytes del PDF o None si falla.
    """
    assert isinstance(form_data, dict), "form_data debe ser dict"

    soffice = _find_soffice()
    if soffice is None:
        print("[excel_pdf_builder] LibreOffice no encontrado. Instalar con: brew install --cask libreoffice")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        # 1. Llenar el Excel
        wb = _fill_workbook(form_data)
        xlsx_path = Path(tmpdir) / "crt_temp.xlsx"
        wb.save(xlsx_path)

        # 2. Convertir con LibreOffice headless
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf",
             "--outdir", tmpdir, str(xlsx_path)],
            capture_output=True, text=True, timeout=30
        )

        if result.returncode != 0:
            print(f"[excel_pdf_builder] Error LibreOffice: {result.stderr}")
            return None

        # 3. Leer el PDF generado
        pdf_path = Path(tmpdir) / "crt_temp.pdf"
        if not pdf_path.exists():
            print(f"[excel_pdf_builder] PDF no generado. stdout: {result.stdout}")
            return None

        return pdf_path.read_bytes()


def save_crt(form_data: dict) -> dict:
    """
    Genera el CRT, guarda el Excel y el PDF en crts_generados/ y devuelve sus rutas.
    Retorna: {"pdf_path": Path | None, "excel_path": Path | None}
    """
    assert isinstance(form_data, dict), "form_data debe ser dict"

    out_dir = Path("crts_generados")
    out_dir.mkdir(exist_ok=True)

    numero = str(form_data.get("f_numero_crt", "crt")).replace("/", "-")

    # Guardar Excel
    excel_path = None
    try:
        wb = _fill_workbook(form_data)
        excel_path = out_dir / f"{numero}.xlsx"
        wb.save(excel_path)
    except Exception as e:
        print(f"[excel_pdf_builder] Error guardando Excel: {e}")

    # Generar PDF
    pdf_path = None
    soffice = _find_soffice()
    if soffice and excel_path and excel_path.exists():
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf",
                 "--outdir", str(out_dir), str(excel_path)],
                capture_output=True, text=True, timeout=30
            )
            candidate = out_dir / f"{numero}.pdf"
            if candidate.exists():
                pdf_path = candidate
            else:
                print(f"[excel_pdf_builder] PDF no encontrado. stdout: {result.stdout} stderr: {result.stderr}")
        except Exception as e:
            print(f"[excel_pdf_builder] Error convirtiendo a PDF: {e}")
    elif not soffice:
        print("[excel_pdf_builder] LibreOffice no encontrado. Instalar con: brew install --cask libreoffice")

    return {"pdf_path": pdf_path, "excel_path": excel_path}
